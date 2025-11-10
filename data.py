# export_categorias.py
import requests
import pandas as pd
from datetime import datetime
from pathlib import Path
import os

import json
from datetime import datetime, date
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



def criar_planilha(dados, nome_arquivo_saida):

    # 1) Preparar caminho/arquivo
    caminho = str(nome_arquivo_saida)
    if not caminho.lower().endswith(".xlsx"):
        caminho += ".xlsx"
    dirpath = os.path.dirname(os.path.abspath(caminho))
    if dirpath and not os.path.exists(dirpath):
        os.makedirs(dirpath, exist_ok=True)

    # 2) Normalizar o payload para (headers, rows)
    headers = None
    rows_matrix = None
    dataset = None

    # Caso especial: dicionário no formato {"headers": [...], " rows": [[...], ...]}
    if isinstance(dados, dict) and "headers" in dados and "rows" in dados and isinstance(dados["rows"], list):
        headers = list(dados["headers"])
        rows_matrix = list(dados["rows"])
    else:
        # Descobrir a lista principal
        if isinstance(dados, dict):
            for key in ("data", "results", "items", "value"):
                if isinstance(dados.get(key), list):
                    dataset = dados[key]
                    break
            if dataset is None:
                dataset = [dados]
        elif isinstance(dados, list):
            dataset = dados
        else:
            dataset = [dados]

        # Transformar dataset em headers + rows
        if all(isinstance(x, dict) for x in dataset):
            # Lista de dicionários: achatar e unificar colunas
            columns = []
            flattened_rows = []
            for item in dataset:
                flat = {}
                stack = [(None, item)]
                while stack:
                    prefix, obj = stack.pop()
                    if isinstance(obj, dict):
                        for k, v in obj.items():
                            key = f"{prefix}.{k}" if prefix else str(k)
                            stack.append((key, v))
                    elif isinstance(obj, (list, tuple)):
                        keyname = prefix if prefix is not None else "lista"
                        flat[keyname] = json.dumps(obj, ensure_ascii=False)
                    else:
                        val = obj
                        if isinstance(obj, (datetime, date)):
                            val = obj.isoformat()
                        elif isinstance(obj, Decimal):
                            try:
                                val = float(obj)
                            except Exception:
                                val = str(obj)
                        keyname = prefix if prefix is not None else "valor"
                        flat[keyname] = val

                # Atualizar ordem de colunas pela primeira aparição
                for k in flat.keys():
                    if k not in columns:
                        columns.append(k)
                flattened_rows.append(flat)

            headers = columns
            rows_matrix = [[row.get(col) for col in columns] for row in flattened_rows]

        elif all(isinstance(x, (list, tuple)) for x in dataset):
            # Lista de listas/tuplas
            max_len = max((len(x) for x in dataset), default=0)
            headers = [f"col_{i+1}" for i in range(max_len)]
            rows_matrix = [list(x) + [None] * (max_len - len(x)) for x in dataset]
        else:
            # Misto ou escalares: uma única coluna
            headers = ["valor"]
            rows_matrix = [[x] for x in dataset]

    # 3) Criar workbook e escrever dados
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    # Cabeçalho
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Linhas + calcular largura de colunas
    col_widths = [len(str(h)) if h is not None else 0 for h in headers]

    for r_idx, row in enumerate(rows_matrix, start=2):
        if len(row) < len(headers):
            row = list(row) + [None] * (len(headers) - len(row))
        for c_idx, val in enumerate(row, start=1):
            cell_val = val
            if isinstance(cell_val, (datetime, date)):
                cell_val = cell_val.isoformat()
            elif isinstance(cell_val, Decimal):
                try:
                    cell_val = float(cell_val)
                except Exception:
                    cell_val = str(cell_val)

            ws.cell(row=r_idx, column=c_idx, value=cell_val)

            tam = len(str(cell_val)) if cell_val is not None else 0
            if tam > col_widths[c_idx - 1]:
                col_widths[c_idx - 1] = tam

    # Autofiltro
    from_col = "A"
    to_col = get_column_letter(len(headers)+1)
    last_row = len(rows_matrix) + 1
    ws.auto_filter.ref = f"{from_col}1:{to_col}{last_row}"

    # Ajuste de largura das colunas
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(w + 2, 60))

    # 4) Salvar e retornar caminho
    wb.save(caminho)
    return caminho


def make_api_request(api_path, method='GET', data=None):
    url = f'https://app.base44.com/api/{api_path}'
    headers = {
        'api_key': 'e77e210c4b674435b2afa613f76ef005',
        'Content-Type': 'application/json'
    }
    if method.upper() == 'GET':
        response = requests.request(method, url, headers=headers, params=data)
    else:
        response = requests.request(method, url, headers=headers, json=data)
    response.raise_for_status()
    return response.json()


entities = make_api_request(f'apps/68f5182879c5fe5a86e409ee/entities/Category')
print(criar_planilha(entities,"Category.xlsx"))

entities = make_api_request(f'apps/68f5182879c5fe5a86e409ee/entities/Transaction')
print(criar_planilha(entities,"Transaction.xlsx"))

entities = make_api_request(f'apps/68f5182879c5fe5a86e409ee/entities/BankAccount')
print(criar_planilha(entities,"BankAccount.xlsx"))

entities = make_api_request(f'apps/68f5182879c5fe5a86e409ee/entities/StatementTransaction')
print(criar_planilha(entities,"StatementTransaction.xlsx"))